import openpyxl
import matplotlib.pyplot as plt
import numpy as np


frequencies = ['50HZ_Bm', '50HZ_Hc', '50HZ_μa', '50HZ_Br', '50HZ_Pcv', '200HZ_Bm', '200HZ_Hc', '200HZ_μa', '200HZ_Br', '200HZ_Pcv', '400HZ_Bm', '400HZ_Hc', '400HZ_μa', '400HZ_Br', '400HZ_Pcv', '800HZ_Bm', '800HZ_Hc', '800HZ_μa', '800HZ_Br', '800HZ_Pcv']
# frequencies = ['50HZ_μa', '200HZ_μa', '400HZ_μa', '800HZ_μa']
# frequencies = ['50HZ_μa']

# 投影方式 (dw_bn/avg/linear)
projection_method = 'dw_bn'

# cls_token 是否打開 (True/False)
cls_token_switch = False

for freq in frequencies:
    # 從Excel文件讀取數據
    # wb = openpyxl.load_workbook(f'bayesian_conv_transformer_par_records_{freq}.xlsx')
    wb = openpyxl.load_workbook(f'Images & Parameters/cvt_records_{freq}_{projection_method}_cls{cls_token_switch}.xlsx')
    ws = wb.active

    # 提取數據並存儲在字典中，以便按列訪問
    headers = ["epoch", "loss", "mae", "val_loss", "val_mae"]
    data = {header: [] for header in headers}

    for row in ws.iter_rows(min_row=2, values_only=True):
        for header, value in zip(headers, row):
            data[header].append(value)

    # 創建圖表
    plt.figure(figsize=(10, 6))

    # 繪制每個物理量的圖線
    epochs = [int(epoch) for epoch in data["epoch"]]  # 將 Epoch 值轉換為整數
    for header in headers[1:]:  # 從第二列開始，忽略Epoch列
        plt.plot(epochs, data[header], label=header)

    # 添加圖例和標簽
    plt.legend()
    plt.xlabel("Epoch")
    plt.ylabel("Value")
    plt.title("Training Metrics Over epochs")

    # 設定 x 軸刻度間隔，這裡設為每 10 個 epoch 顯示一個刻度
    plt.xticks(np.arange(min(epochs)-1, max(epochs), 50))

    # # 設定 X 軸範圍
    # plt.xlim(0, 200)  # 将Y轴范围限制在0到100之间

    # # 設定 Y 軸範圍
    # plt.ylim(0, 2000)  # 将Y轴范围限制在0到100之间

    # 假設你所有數據點都存儲在一個list中
    all_values = []
    for header in headers[1:]:
        all_values.extend(data[header])

    # 計算95%的值作為Y軸的上限
    upper_limit = np.percentile(all_values, 99.95)

    plt.ylim([0, upper_limit])

    # 顯示圖表
    plt.grid(True)
    plt.savefig(f'Images & Parameters/Training Metrics_{freq}_{projection_method}_cls{cls_token_switch}.png')  # 儲存圖片
    # plt.show()
    plt.clf() # 清空畫布（plt.show()內建有）
