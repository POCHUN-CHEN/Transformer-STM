import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# def process_data(group):
#     # 計算平均值和標準差
#     mean = group.mean()
#     std = group.std()
#     # 處理超過標準差的數據
#     outliers = (group < (mean - std)) | (group > (mean + std))
#     group[outliers] = None  # 例如，可以將異常值設為None
#     return group

def process_data(group):
    # 計算Q1和Q3
    Q1 = group.quantile(0.25)
    Q3 = group.quantile(0.75)
    # 計算IQR
    IQR = Q3 - Q1
    # 定義異常值的範圍
    outliers = (group < (Q1 - 1.5 * IQR)) | (group > (Q3 + 1.5 * IQR))
    # 處理異常值，這裡可以根據需要來決定如何處理
    group[outliers] = None  # 例如，可以將異常值設為None
    return group

def apply_styles(workbook, sheet_name):
    # 定義底色和字體樣式
    fill_color1 = PatternFill(start_color="FFFF93", end_color="FFFF93", fill_type="solid")
    fill_color2 = PatternFill(start_color="CCFF80", end_color="CCFF80", fill_type="solid")
    font_style = Font(name="Arial", size=10)
    
    sheet = workbook[sheet_name]
    for row in range(1, sheet.max_row + 1):
        # 跳過第一行的底色設定，但仍設定字體樣式
        fill_color = fill_color1 if (row - 2) // 5 % 2 == 0 else fill_color2
        for cell in sheet[row]:
            cell.font = font_style
            if row != 1:  # 第一行不設定底色
                cell.fill = fill_color

def process_excel(file_path, output_path):
    data = pd.read_excel(file_path, index_col=None) # 明確指定不將任何列作為索引
    # 保持第一列不變，處理其他數據
    index_col = data.columns[0]
    index_values = data[index_col]  # 保留Col A的值
    processed_data = data.copy()
    # processed_data = data.iloc[:, 1:].groupby(data.index // 5).apply(lambda x: x.apply(process_data))
    # processed_data.insert(0, index_col, data[index_col])
    processed_data.iloc[:, 1:] = data.iloc[:, 1:].groupby(data.index // 5).apply(lambda x: x.apply(process_data)).reset_index(drop=True)
    processed_data[index_col] = index_values  # 恢復Col A的值
    
    # 將處理後的數據存儲為Excel檔案
    processed_data.to_excel(output_path, index=False)
    
    # 讀取剛存儲的Excel檔案並應用樣式
    workbook = load_workbook(output_path)
    apply_styles(workbook, workbook.sheetnames[0])

    # 設定Col A的第一行為空白
    sheet = workbook[workbook.sheetnames[0]]
    sheet.cell(row=1, column=1).value = ""

    workbook.save(output_path)
    
    print(f'Processed file with styles applied saved.')

if __name__ == "__main__":
    # 獲取當前腳本所在的目錄
    script_dir = os.path.dirname(os.path.abspath(__file__))

    # 指定原始和處理後的檔案路徑
    file_path = os.path.join(script_dir, '../Excel/Circle_test.xlsx')
    output_path = os.path.join(script_dir, '../Excel/Processed_Circle_test.xlsx')

    # 執行處理函數
    process_excel(file_path, output_path)
